# =====================================================================
# 사전 설치: pip install requests numpy pandas matplotlib pillow openpyxl
# =====================================================================

# ---------------------------------------------------------------------
# [1] requests : 인터넷에서 데이터 가져오기 (HTTP GET 요청)
# ---------------------------------------------------------------------
def demo_requests():
    # pip install requests
    import requests     # 웹사이트에 접속하거나, API에 데이터 요청하는 라이브러리
    import urllib3

    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning) # SSL 인증서 검증 비활성 및 경고 숨김

    # GET방식으로 서버(catfact.ninja/fact, 고양이 정보 공개 API)에 데이터 요청
    response = requests.get("https://catfact.ninja/fact", verify=False)
    
    cats = response.json()                      # JSON 데이터를 파이썬 딕셔너리로 변환.

    print(cats)                                 # 전체 출력
    print(f"{cats['fact']}\n{cats['length']}")  # key를 이용해서 value 출력

    for cat in cats:                            # 딕셔너리를 for문으로 반복
        print(f"{cat}: {cats[cat]}")


# ---------------------------------------------------------------------
# [2] numpy : 수치/배열 계산
#     소재 - 우리 반 학생 5명의 중간고사 점수 통계 내기
# ---------------------------------------------------------------------
def demo_numpy():
    # pip install numpy
    # numpy : 빠른 수치 계산과 배열(행렬) 연산을 위한 핵심 라이브러리.
    import numpy as np
    
    scores = np.array([88, 92, 75, 100, 67])        # 학생 5명의 점수를 numpy 배열(ndarray)로 생성

    # 일반 리스트와 달리 배열은 한 번에 사칙연산이 가능.
    bonus_scores = scores + 3           # 모든 점수에 보너스 +3점을 한 번에 더하기

    print("[numpy] 원점수:", scores)
    print("[numpy] +3 보너스:", bonus_scores)
    print(f"[numpy] 평균: {scores.mean():.2f}점")       # numpy 배열의 메서드 사용
    print(f"[numpy] 최고점: {scores.max()}점 / 최저점: {scores.min()}점")
    print(f"[numpy] 표준편차: {scores.std():.2f}")

# ---------------------------------------------------------------------
# [3] pandas : 표(테이블) 형태의 데이터 분석
#     소재 - K-POP 아이돌 그룹의 멤버 정보 데이터프레임
# ---------------------------------------------------------------------
def demo_pandas():
    # pip install pandas
    # pandas : 엑셀처럼 행과 열로 된 데이터를 다룰 때 쓰는 라이브러리
    import pandas as pd

    idol_data = {
        "이름":   ["진",  "RM",  "슈가", "제이홉", "지민", "뷔",  "정국"],
        "포지션": ["보컬","리더","래퍼","래퍼",  "댄서","보컬","메인보컬"],
        "나이":   [33,    31,    32,    32,     30,    30,    28]
    }
    
    # 딕셔너리를 DataFrame(표)으로 변환
    df = pd.DataFrame(idol_data)    # key : 열(column) 이름, value : 각 열의 데이터로 변환

    print("[pandas] 전체 데이터:")
    print(df)

    print(f"\n[pandas] 평균 나이: {df['나이'].mean():.1f}세")   # 평균 나이 계산
    
    print("\n[pandas] 30세 이상 멤버:")
    print(df[df["나이"] >= 30])             # 조건 필터링 - 30세 이상 멤버만 추출

# ---------------------------------------------------------------------
# [4] matplotlib : 그래프 그리기 (한글 폰트 + 안전 저장 버전)
# ---------------------------------------------------------------------
def demo_matplotlib():
    # pip install matplotlib
    import matplotlib.pyplot as plt     # 실제 그래프 그리는 기능
    import matplotlib                   # 전체 설정 변경, 한글 폰트 설정에 사용

    matplotlib.rc("font", family="Malgun Gothic")           # 한글 폰트 사용시
    matplotlib.rcParams["axes.unicode_minus"] = False       # 한글 폰트 사용 시 음수 부호(-)가 깨지는 현상 방지

    # ── 데이터 준비 ──────────────────────────────────────────────
    days  = ["월", "화", "수", "목", "금", "토", "일"]      # X축 데이터
    hours = [1.5, 2.0, 1.0, 2.5, 3.0, 5.5, 4.0]             # Y축 데이터

    # ── 그래프 그리기 ────────────────────────────────────────────
    plt.figure(figsize=(7, 4))                          # 새로운 그래프 화면 생성, 가로, 세로 비율
    plt.plot(days, hours, marker="o", color="tomato", linewidth=2)  # marker="o":각 데이터 위치에 동그라미 표시
                                            # color="tomato" : 선의 색을 토마토색 계열로 지정
                                            # linewidth=2 : 선의 두께를 2로 설정
    plt.title("이번 주 게임 플레이 시간")
    plt.xlabel("요일")
    plt.ylabel("시간(h)")
    plt.grid(True)
    plt.show()                              # 화면에 띄우기
    plt.close()                             # 메모리에서 figure 정리

# ---------------------------------------------------------------------
# [5] Pillow(PIL) : 이미지 처리
#     소재 - 새 이미지를 만들어서 텍스트 넣고, 회전/흑백 변환하기
# ---------------------------------------------------------------------
def demo_pillow():
    # pip install pillow
    from PIL import Image, ImageDraw        # 새 이미지 생성
    from pathlib import Path
    
    # ── 저장 경로 준비 ────────────────────────────────────────────
    try:
        base_dir = Path(__file__).resolve().parent  # 현재 파일이 있는 절대 경로 구하기
    except NameError:
        base_dir = Path.cwd()

    save_dir = base_dir / "pillow_lesson_result"    # 현재 위치에 폴더 생성
    save_dir.mkdir(exist_ok=True)                   # 폴더 이미 존재하면 패스

    original_path = save_dir / "pillow_original.png"    # 저장할 경로 및 파일 생성
    rotated_path = save_dir / "pillow_rotated.png"
    gray_path = save_dir / "pillow_gray.png"

    # ── 그림 그리기 ────────────────────────────────────────────
    img = Image.new("RGB", (300, 200), color=(135, 206, 250))   # 가로 300픽셀, 세로 200픽셀 크기의 하늘색 이미지 생성

    # 이미지 그리기
    draw = ImageDraw.Draw(img)          # 이미지 위에 그림을 그릴 수 있는 도구 만들기
    draw.rectangle([(20, 20), (280, 180)], outline="white", width=3)    # 사각형 그리기(20, 20)은 왼쪽 위, (280, 180)은 오른쪽 아래
    draw.text((40, 90), "Hello, COS Pro!", fill="white")    # 이미지 위에 글자 쓰기, (40, 90) : 글자가 시작되는 위치.
    rotated = img.rotate(15, expand=True)     # 이미지를 15도 회전, expand=True : 회전할 때 이미지가 잘리지 않도록 캔버스 크기 확장
    gray = img.convert("L")                         # 이미지를 흑백으로 변환

    # ── 이미지 파일 저장하기 ────────────────────────────────────────────
    img.save(original_path)
    rotated.save(rotated_path)
    gray.save(gray_path)

    print("[Pillow] 원본/회전/흑백 이미지 3개 저장 완료")
    print("원본 이미지:", original_path)
    print("회전 이미지:", rotated_path)
    print("흑백 이미지:", gray_path)

# ---------------------------------------------------------------------
# [6] openpyxl : 엑셀 파일 만들기/읽기
#     소재 - 내 용돈 가계부를 엑셀로 저장하기
# ---------------------------------------------------------------------
def demo_openpyxl():
    # pip install openpyxl
    from openpyxl import Workbook, load_workbook    # 엑셀(.xlsx) 파일을 만들고 읽는 라이브러리
    from pathlib import Path
        
    # 1) 새 워크북(엑셀 파일) 생성
    wb = Workbook()
    ws = wb.active            # 현재 활성화된 시트(첫 시트) 선택
    ws.title = "용돈가계부"    # 시트 이름 변경

    # 2) 머리글(첫 줄) 작성
    ws.append(["날짜", "항목", "금액"])

    # 3) 데이터 한 줄씩 추가
    items = [
        ("5/01", "편의점 간식",   -3500),
        ("5/03", "용돈",          +30000),
        ("5/05", "버스카드 충전", -10000),
        ("5/10", "친구 생일선물", -8000),
    ]
    for row in items:
        ws.append(row)

    # 4) 합계 수식 추가 - 엑셀 함수도 그대로 사용 가능
    ws["A6"] = "합계"
    ws["C6"] = "=SUM(C2:C5)"

    # 5) 파일로 저장
    base_dir = Path(__file__).resolve().parent          # 현재 파일이 있는 절대 경로 구하기
    save_dir = base_dir / "openpyxl_lesson_result"      # 현재 위치에 폴더 생성
    save_dir.mkdir(exist_ok=True)                       # 폴더 이미 존재하면 패스
    
    path = save_dir / "pocket_mony.xlsx"
    wb.save(path)
    print("[openpyxl] 'pocket_money.xlsx' 저장 완료")

    # 6) 다시 읽어서 내용 확인
    wb2 = load_workbook(path)
    ws2 = wb2.active
    print("[openpyxl] 저장된 내용 확인:")
    for row in ws2.iter_rows(values_only=True):
        print("  ", row)

# ---------------------------------------------------------------------
# main : 통합 실행 관리자
#   - 원하는 데모만 골라서 호출할 수 있도록 dict로 관리
# ---------------------------------------------------------------------
def main():
    demos = {
        "1": ("requests   - 고양이 정보 가져오기", demo_requests),
        "2": ("numpy      - 시험 점수 통계",       demo_numpy),
        "3": ("pandas     - K-POP 데이터분석",     demo_pandas),
        "4": ("matplotlib - 게임시간 그래프",      demo_matplotlib),
        "5": ("Pillow     - 이미지 만들기",        demo_pillow),
        "6": ("openpyxl   - 엑셀 가계부",          demo_openpyxl),
    }

    print("=" * 50)
    print(" COS Pro 서드파티 모듈 실습")
    print("=" * 50)
    for key, (desc, _) in demos.items():
        print(f"  {key}. {desc}")
    print("=" * 50)

    # ── 학생 실습 옵션 ────────────────────────────────────────────────
    # (A) 전체 실행: 아래 for 루프 그대로 사용
    # (B) 개별 실행: 아래 줄 주석 해제 → 번호 입력해서 한 개만 실행
    # choice = input("실행할 번호를 입력하세요: ")
    # demos[choice][1](); return
    # ─────────────────────────────────────────────────────────────────

    for key, (desc, func) in demos.items():
        print(f"\n>>> [{key}] {desc} 실행")
        try:
            func()
        except Exception as e:
            print(f"  (오류) {e}")


if __name__ == "__main__":
    main()
